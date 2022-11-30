#the purpose from this  this file is to create a copy from the last bd_interm.xlsx file before editing it
import warnings
warnings.filterwarnings("ignore")

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import pandas as pd
from constant import bd_interm_hist,bd_interm
# checking how many sheets in the workbook and returning the number of sheets checked
def checking_sheet_names(x = bd_interm_hist):
    wb = load_workbook(x)
    return len(wb.sheetnames)

def adding_new_copy(df, x = bd_interm_hist):
    sheet_len = checking_sheet_names()
    newSheetName = 'sheet' +str(sheet_len+1)

    # load in existing workbook
    wb = load_workbook(x)
    # create new sheet with name assigned by title
    wb.create_sheet(newSheetName)
    # find which sheet is the new sheet
    newSheetIndex = len(wb.sheetnames) - 1
    # set the newest sheet as the active sheet
    wb.active = newSheetIndex
    worksheet = wb.active
    # Insert Dataframe
    for r in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(r)
    wb.save(x)

